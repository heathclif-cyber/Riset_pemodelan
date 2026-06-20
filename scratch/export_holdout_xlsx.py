# -*- coding: utf-8 -*-
"""Export holdout diagnostic lengkap ke Excel."""
from __future__ import annotations

import json
import sys
from datetime import datetime
from pathlib import Path

import joblib
import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows

ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT))

from importlib import import_module
from core.evaluator import full_trading_report
from core.models import load_lstm
from pipeline.ic32_fusion_shared import build_per_bar_thresholds, load_b_dir_hmm_cfg
from pipeline.backtest_utils import hierarchical_predict, compute_guardian_static_array
from config import (
    ALL_COINS, HOLDOUT_DIR, MODEL_DIR, LABEL_MAP,
    MODAL_PER_TRADE, LEVERAGE_SIM, FEE_PER_SIDE, SLIPPAGE_PER_SIDE,
    MAX_HOLDING_BARS, GUARDIAN_MIN_HOLD_BARS,
    LIVE_MAX_OPEN_POSITIONS, LIVE_DAILY_LOSS_LIMIT,
    SWING_LABEL_MIN_RR, SWING_LABEL_MIN_TP, SWING_LABEL_MAX_SL,
    TP_SL_FALLBACK_TP, TP_SL_FALLBACK_SL,
    TRAILING_STOP_ENABLED, TRAILING_STOP_ATR, TRAILING_STOP_MIN_BARS,
    OOS_START, TP_SL_TRIGGER_MODE,
)

h07 = import_module("pipeline.07h_holdout_ic32_scale_in_diag")
_apply_live_config = h07._apply_live_config
_run_holdout = h07._run_holdout
# Satu-satunya variant = setup live VPS (scale_in max 2/koin)
LIVE_VARIANT = {
    "label": "live_current",
    "enabled": True,
    "max_per_coin": 2,
    "exit_mode": "scale_in",
}
RUN_DIR = MODEL_DIR / "runs" / "ic32_regime_v1"
OUT_XLSX = Path(r"D:\Datatrade_ic32regime.xlsx")
INF_CFG = MODEL_DIR / "inference_config.json"
WITA = __import__("datetime").timezone(__import__("datetime").timedelta(hours=8))


def _pf(gw: float, gl: float):
    if gl <= 0:
        return None if gw <= 0 else 999.0
    return gw / gl


def _collect_live_trades() -> list:
    live_cfg = _apply_live_config()
    hmm_cfg = load_b_dir_hmm_cfg()
    gdn = h07._load_guardian_cont()
    with open(MODEL_DIR / "feature_cols_ic32_regime.json", encoding="utf-8") as f:
        feat_cols = json.load(f)
    with open(MODEL_DIR / "feature_cols_lstm_temporal.json", encoding="utf-8") as f:
        lstm_feats = json.load(f)[:11]
    lgbm = joblib.load(RUN_DIR / "lgbm.pkl")
    lstm = load_lstm(MODEL_DIR / "lstm_best.pt", device="cpu")
    lstm_scaler = joblib.load(MODEL_DIR / "lstm_scaler.pkl")

    trades = []
    for sym in ALL_COINS:
        trades.extend(
            _run_holdout(sym, hmm_cfg, live_cfg, gdn, feat_cols, lstm_feats,
                         lgbm, lstm, lstm_scaler, LIVE_VARIANT)
        )
    return trades


def _ts_str(val) -> str | None:
    if val is None:
        return None
    ts = pd.Timestamp(val)
    if ts.tzinfo is not None:
        ts = ts.tz_convert("UTC").tz_localize(None)
    return ts.strftime("%Y-%m-%d %H:%M:%S")


def _trades_df(trades: list, variant: str) -> pd.DataFrame:
    if not trades:
        return pd.DataFrame()
    rows = []
    for t in trades:
        rows.append({
            "variant": variant,
            "date": pd.Timestamp(t["ts_in"]).date() if t.get("ts_in") else None,
            "ts_in": _ts_str(t.get("ts_in")),
            "ts_out": _ts_str(t.get("ts_out")),
            "symbol": t.get("symbol"),
            "direction": t.get("direction"),
            "entry": t.get("entry"),
            "exit": t.get("exit"),
            "tp": t.get("tp"),
            "sl": t.get("sl"),
            "outcome": t.get("outcome"),
            "net_pnl": t.get("net_pnl"),
            "modal_used": t.get("modal_used", MODAL_PER_TRADE),
            "n_legs": t.get("n_legs", 1),
            "scale_in": t.get("scale_in", False),
            "hold_bars": t.get("bar_out", 0) - t.get("bar_in", 0),
            "is_win": t.get("net_pnl", 0) > 0,
        })
    df = pd.DataFrame(rows)
    df["net_pnl"] = df["net_pnl"].round(4)
    return df.sort_values(["date", "symbol", "ts_in"]).reset_index(drop=True)


def _daily_all(df: pd.DataFrame) -> pd.DataFrame:
    daily = []
    for date, sub in df.groupby("date"):
        wins = sub[sub["is_win"]]
        losses = sub[~sub["is_win"]]
        gw = float(wins["net_pnl"].sum())
        gl = abs(float(losses["net_pnl"].sum()))
        pnl = float(sub["net_pnl"].sum())
        coins = sub["symbol"].nunique()
        long_n = int((sub["direction"] == "LONG").sum())
        short_n = int((sub["direction"] == "SHORT").sum())
        daily.append({
            "date": date,
            "trades": len(sub),
            "wins": len(wins),
            "losses": len(losses),
            "win_rate_pct": round(len(wins) / len(sub) * 100, 1) if len(sub) else 0,
            "profit_factor": round(_pf(gw, gl), 3) if _pf(gw, gl) is not None else None,
            "gross_win": round(gw, 2),
            "gross_loss": round(-gl, 2),
            "net_pnl": round(pnl, 2),
            "coins_traded": coins,
            "long_trades": long_n,
            "short_trades": short_n,
            "avg_pnl_per_trade": round(pnl / len(sub), 4) if len(sub) else 0,
            "day_type": "good" if pnl > 0 else ("flat" if pnl == 0 else "bad"),
        })
    out = pd.DataFrame(daily).sort_values("date").reset_index(drop=True)
    out["cum_pnl"] = out["net_pnl"].cumsum().round(2)
    return out


def _daily_by_coin(df: pd.DataFrame) -> pd.DataFrame:
    rows = []
    for (date, sym), sub in df.groupby(["date", "symbol"]):
        wins = sub[sub["is_win"]]
        losses = sub[~sub["is_win"]]
        gw = float(wins["net_pnl"].sum())
        gl = abs(float(losses["net_pnl"].sum()))
        pnl = float(sub["net_pnl"].sum())
        rows.append({
            "date": date,
            "symbol": sym,
            "trades": len(sub),
            "wins": len(wins),
            "losses": len(losses),
            "win_rate_pct": round(len(wins) / len(sub) * 100, 1) if len(sub) else 0,
            "profit_factor": round(_pf(gw, gl), 3) if _pf(gw, gl) is not None else None,
            "gross_win": round(gw, 2),
            "gross_loss": round(-gl, 2),
            "net_pnl": round(pnl, 2),
            "long_trades": int((sub["direction"] == "LONG").sum()),
            "short_trades": int((sub["direction"] == "SHORT").sum()),
            "avg_modal": round(float(sub["modal_used"].mean()), 2),
            "pct_2leg": round((sub["n_legs"] >= 2).mean() * 100, 1),
        })
    return pd.DataFrame(rows).sort_values(["date", "symbol"]).reset_index(drop=True)


def _coin_summary(df: pd.DataFrame) -> pd.DataFrame:
    rows = []
    for sym, sub in df.groupby("symbol"):
        wins = sub[sub["is_win"]]
        losses = sub[~sub["is_win"]]
        gw = float(wins["net_pnl"].sum())
        gl = abs(float(losses["net_pnl"].sum()))
        pnl = float(sub["net_pnl"].sum())
        rows.append({
            "symbol": sym,
            "trades": len(sub),
            "wins": len(wins),
            "losses": len(losses),
            "win_rate_pct": round(len(wins) / len(sub) * 100, 1),
            "profit_factor": round(_pf(gw, gl), 3) if _pf(gw, gl) is not None else None,
            "net_pnl": round(pnl, 2),
            "pnl_per_trade": round(pnl / len(sub), 4),
            "long_trades": int((sub["direction"] == "LONG").sum()),
            "short_trades": int((sub["direction"] == "SHORT").sum()),
            "avg_modal": round(float(sub["modal_used"].mean()), 2),
            "pct_2leg": round((sub["n_legs"] >= 2).mean() * 100, 1),
            "trading_days": int(sub["date"].nunique()),
        })
    return pd.DataFrame(rows).sort_values("net_pnl", ascending=False).reset_index(drop=True)


def _summary_row(df: pd.DataFrame) -> pd.DataFrame:
    if df.empty:
        return pd.DataFrame()
    wins = df[df["is_win"]]
    losses = df[~df["is_win"]]
    gw = float(wins["net_pnl"].sum())
    gl = abs(float(losses["net_pnl"].sum()))
    pnl = float(df["net_pnl"].sum())
    daily = _daily_all(df)
    return pd.DataFrame([{
        "setup": "live_current",
        "period_start": str(df["date"].min()),
        "period_end": str(df["date"].max()),
        "trading_days": int(df["date"].nunique()),
        "total_trades": len(df),
        "trades_per_day": round(len(df) / df["date"].nunique(), 2),
        "wins": len(wins),
        "losses": len(losses),
        "win_rate_pct": round(len(wins) / len(df) * 100, 2),
        "profit_factor": round(_pf(gw, gl), 3),
        "total_pnl": round(pnl, 2),
        "pnl_per_trade": round(pnl / len(df), 4),
        "pnl_per_day": round(pnl / df["date"].nunique(), 2),
        "good_days": int((daily["net_pnl"] > 0).sum()),
        "bad_days": int((daily["net_pnl"] < 0).sum()),
        "max_loss_streak_days": _max_bad_streak(daily),
    }])


def _load_live_setup() -> dict:
    with open(INF_CFG, encoding="utf-8") as f:
        cfg = json.load(f)
    risk = cfg.get("risk", {})
    pyr = cfg.get("pyramiding", {})
    gdn = cfg.get("guardian", {})
    rr = cfg.get("rr_gate", {})
    fe = cfg.get("feature_engineering", {})
    hmm = cfg.get("hmm", {}).get("per_state_thresholds", {})
    return {
        "model_version": cfg.get("model_version"),
        "cascade_mode": cfg.get("cascade", {}).get("mode"),
        "conf_entry": cfg.get("cascade", {}).get("confidence_threshold_entry"),
        "guardian_exit": gdn.get("exit_threshold"),
        "guardian_min_hold": gdn.get("min_hold_bars"),
        "guardian_model": gdn.get("model_file"),
        "sl_trigger_mode": rr.get("sl_trigger_mode"),
        "positioning_mode": fe.get("positioning_mode", "training_parity"),
        "modal_per_trade": risk.get("modal_per_trade"),
        "leverage": risk.get("leverage_recommended"),
        "max_open_positions": risk.get("max_open_positions", LIVE_MAX_OPEN_POSITIONS),
        "daily_loss_limit": risk.get("daily_loss_limit", LIVE_DAILY_LOSS_LIMIT),
        "pyramiding_enabled": pyr.get("enabled"),
        "pyramiding_exit_mode": pyr.get("exit_mode"),
        "pyramiding_max_per_coin": pyr.get("max_positions_per_coin"),
        "hmm_thresholds": json.dumps(hmm),
    }


def _wita_date(ts: pd.Timestamp) -> str:
    if ts.tzinfo is None:
        ts = ts.tz_localize("UTC")
    return ts.tz_convert(WITA).date().isoformat()


def _apply_risk_gates(df: pd.DataFrame, max_pos: int, daily_lim: int) -> pd.DataFrame:
    """Filter trade rows kronologis — mirror execution.py live."""
    if df.empty:
        return df
    work = df.copy()
    work["ts_in_dt"] = pd.to_datetime(work["ts_in"])
    work["ts_out_dt"] = pd.to_datetime(work["ts_out"])
    work = work.sort_values("ts_in_dt").reset_index(drop=True)
    work["trade_id"] = np.arange(len(work))
    lookup = work.set_index("trade_id").to_dict("index")

    events = []
    for _, r in work.iterrows():
        tid = int(r["trade_id"])
        events.append((r["ts_in_dt"], "entry", tid))
        events.append((r["ts_out_dt"], "exit", tid))
    events.sort(key=lambda x: (x[0], 0 if x[1] == "exit" else 1))

    open_ids: set[int] = set()
    daily_losses: dict[str, int] = {}
    accepted: set[int] = set()

    for ts, kind, tid in events:
        if kind == "exit":
            if tid in open_ids:
                open_ids.discard(tid)
                pnl = lookup[tid]["net_pnl"]
                if pnl < 0:
                    day = _wita_date(ts)
                    daily_losses[day] = daily_losses.get(day, 0) + 1
        else:
            day = _wita_date(ts)
            if len(open_ids) >= max_pos:
                continue
            if daily_losses.get(day, 0) >= daily_lim:
                continue
            open_ids.add(tid)
            accepted.add(tid)

    out = work[work["trade_id"].isin(accepted)].drop(columns=["ts_in_dt", "ts_out_dt", "trade_id"])
    out["variant"] = "live_current"
    return out.reset_index(drop=True)


def _live_setup_rows(setup: dict) -> pd.DataFrame:
    rows = [
        ("Setup", "Deploy VPS saat ini (scale_in + risk gate)"),
        ("Holdout period", f"{OOS_START.date()} s/d Jun 2026"),
        ("", ""),
        ("--- Model ---", ""),
        ("model_version", setup["model_version"]),
        ("cascade_mode", setup["cascade_mode"]),
        ("conf_entry", setup["conf_entry"]),
        ("guardian", f"{setup['guardian_model']} exit={setup['guardian_exit']} min_hold={setup['guardian_min_hold']}"),
        ("sl_trigger_mode", setup["sl_trigger_mode"]),
        ("positioning_mode", setup["positioning_mode"]),
        ("hmm_per_state_thr", setup["hmm_thresholds"]),
        ("", ""),
        ("--- Risk (live VPS) ---", ""),
        ("modal_per_trade", f"${setup['modal_per_trade']}"),
        ("leverage", f"{setup['leverage']}x"),
        ("max_open_positions", setup["max_open_positions"]),
        ("daily_loss_limit", setup["daily_loss_limit"]),
        ("", ""),
        ("--- Pyramiding (live) ---", ""),
        ("enabled", setup["pyramiding_enabled"]),
        ("exit_mode", setup["pyramiding_exit_mode"]),
        ("max_positions_per_coin", setup["pyramiding_max_per_coin"]),
        ("", ""),
        ("--- Catatan ---", ""),
        ("vs live bug period", "Holdout pakai fitur benar (training_parity); live bug LSR=0 tidak termasuk"),
        ("retrain", "Tidak perlu — ini evaluasi stack deploy saat ini"),
    ]
    return pd.DataFrame(rows, columns=["parameter", "value"])


def _max_bad_streak(daily: pd.DataFrame) -> int:
    m = cur = 0
    for pnl in daily.sort_values("date")["net_pnl"]:
        if pnl < 0:
            cur += 1
            m = max(m, cur)
        else:
            cur = 0
    return m


def _style_header(ws, row=1):
    fill = PatternFill("solid", fgColor="1F4E79")
    font = Font(bold=True, color="FFFFFF", name="Arial")
    for cell in ws[row]:
        if cell.value:
            cell.fill = fill
            cell.font = font
            cell.alignment = Alignment(horizontal="center", vertical="center")


def _write_df(ws, df: pd.DataFrame, start_row=1):
    for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), start_row):
        for c_idx, val in enumerate(row, 1):
            ws.cell(row=r_idx, column=c_idx, value=val)
    _style_header(ws, start_row)
    for col in range(1, len(df.columns) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 14


def _add_readme(wb: Workbook, meta: dict, setup: dict):
    ws = wb.active
    ws.title = "Info"
    lines = [
        ["IC32 Regime Holdout — Setup Live Saat Ini"],
        [""],
        ["Generated", meta["created"]],
        ["Stack", meta["stack"]],
        ["Holdout start", str(OOS_START.date())],
        ["Modal per trade", f"${setup['modal_per_trade']}"],
        ["Leverage", f"{setup['leverage']}x"],
        ["Max open positions", setup["max_open_positions"]],
        ["Daily loss limit", setup["daily_loss_limit"]],
        ["Pyramiding", f"{setup['pyramiding_exit_mode']} max {setup['pyramiding_max_per_coin']}/coin"],
        ["SL trigger", setup["sl_trigger_mode"]],
        ["Positioning", setup["positioning_mode"]],
        ["Coins", len(ALL_COINS)],
        [""],
        ["Sheets"],
        ["Live_Setup", "Parameter deploy VPS"],
        ["Summary", "Metrik holdout setup live"],
        ["Trades", "Detail trade (setelah risk gate)"],
        ["Daily", "Agregat harian"],
        ["DailyByCoin", "Breakdown per koin per hari"],
        ["CoinSummary", "Ringkasan per koin"],
        [""],
        ["Source", "scratch/export_holdout_xlsx.py"],
        ["Methodology", "holdout Apr-Jun 2026, B-dir + continuation_v1 + min_hold=4"],
    ]
    for i, row in enumerate(lines, 1):
        ws.cell(row=i, column=1, value=row[0])
        if len(row) > 1:
            ws.cell(row=i, column=2, value=row[1])
    ws["A1"].font = Font(bold=True, size=14, name="Arial")
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 55


def main():
    setup = _load_live_setup()
    print(f"Collecting holdout trades @ ${MODAL_PER_TRADE} (setup live saja)...")
    raw = _collect_live_trades()
    scale = _trades_df(raw, "live_current")
    df = _apply_risk_gates(
        scale, int(setup["max_open_positions"]), int(setup["daily_loss_limit"]),
    )

    wb = Workbook()
    meta = {
        "created": datetime.now().isoformat(timespec="seconds"),
        "stack": "B-dir + continuation_v1 + min_hold=4 + SL close + scale_in + risk gate",
    }
    _add_readme(wb, meta, setup)

    ws_setup = wb.create_sheet("Live_Setup")
    _write_df(ws_setup, _live_setup_rows(setup))

    ws_sum = wb.create_sheet("Summary")
    _write_df(ws_sum, _summary_row(df))

    _write_df(wb.create_sheet("Daily"), _daily_all(df))
    _write_df(wb.create_sheet("Trades"), df)
    _write_df(wb.create_sheet("DailyByCoin"), _daily_by_coin(df))
    _write_df(wb.create_sheet("CoinSummary"), _coin_summary(df))

    out_path = OUT_XLSX
    try:
        wb.save(out_path)
    except PermissionError:
        out_path = OUT_XLSX.with_name(OUT_XLSX.stem + "_new.xlsx")
        wb.save(out_path)
        print(f"[WARN] {OUT_XLSX} locked - saved to {out_path}")
    print(f"Saved {out_path}")
    print(f"Sheets: {wb.sheetnames}")
    if not df.empty:
        wins = df[df["is_win"]]
        losses = df[~df["is_win"]]
        gw, gl = float(wins["net_pnl"].sum()), abs(float(losses["net_pnl"].sum()))
        pf = _pf(gw, gl)
        pnl = float(df["net_pnl"].sum())
        pf_s = f"{pf:.2f}" if pf is not None else "n/a"
        print(
            f"  live_current: {len(df)} trades, WR={100*len(wins)/len(df):.1f}%, "
            f"PF={pf_s}, PnL=${pnl:+.2f}, PPT=${pnl/len(df):+.4f}"
        )


if __name__ == "__main__":
    main()